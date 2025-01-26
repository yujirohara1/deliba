from flask import Flask, render_template, g, request, redirect, url_for, Response, abort, session, jsonify, make_response, send_file
from hamlish_jinja import HamlishExtension
from werkzeug.datastructures import ImmutableDict
import os
from flask_login import LoginManager, login_user, logout_user, login_required, UserMixin, current_user
from collections import defaultdict
from datetime import timedelta
import datetime
import pytz
from flask_bootstrap import Bootstrap
from marshmallow_sqlalchemy import ModelSchema
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.pagesizes import A4, portrait
from reportlab.platypus import Table, TableStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from api.database import db, ma
from models.item import Item, ItemSchema, VItemGroup, VItemGroupSchema
from models.orderitem import OrderItem, OrderItemSchema, VOrderItem, VOrderItemSchema, VOrderedGroup, VOrderedGroupSchema
from models.customer import Customer, CustomerSchema, CustomerNentuki, CustomerNentukiSchema
from models.mstsetting import MstSetting, MstSettingSchema
from models.daicho import Daicho, DaichoSchema, VDaichoA, VDaichoASchema
from models.seikyu import Seikyu, SeikyuSchema, VSeikyuA, VSeikyuASchema, VSeikyuB, VSeikyuBSchema, VSeikyuC, VSeikyuCSchema
from models.kakute import Kakute, KakuteSchema
from print.print_seikyu import *
from sqlalchemy.sql import text
from sqlalchemy import distinct
from sqlalchemy import asc
import json
# from rq import Queue
# from worker import conn
import PyPDF2
# from bottle import route, run
import smtplib
from email.mime.text import MIMEText
from email.utils import formatdate
import csv
import shutil
import openpyxl
from openpyxl.worksheet.pagebreak import Break 
# import logging 
# logging.basicConfig()
# logging.getLogger('sqlalchemy.engine').setLevel(logging.INFO)


DELIMIT = "@|@|@"
XLSX_MIMETYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

class FlaskWithHamlish(Flask):
    jinja_options = ImmutableDict(
        extensions=[HamlishExtension]
    )
app = FlaskWithHamlish(__name__)
bootstrap = Bootstrap(app)


login_manager = LoginManager()
login_manager.init_app(app)
app.config['SECRET_KEY'] = "secret"
mail_address = os.environ.get('MAIL_ADDR')
mail_password = os.environ.get('MAIL_PASS')

class User(UserMixin):
    def __init__(self, id, name, password, tenant_id):
        self.id = id
        self.name = name
        self.password = password
        self.tenant_id = tenant_id

# ログイン用ユーザー作成
users = {
    1: User(1, "yujiro", "yjrhr1102", "demo"),
    2: User(2, "seiya", "seiya7293", "hara"),
    3: User(3, "yasu", "3021", "sato"),
    4: User(4, "seiya2", "seiya7294", "sato"),
    5: User(5, "setu", "0301", "hoiku"),
    6: User(6, "seiya3", "seiya7295", "hoiku"),
    11: User(11, "aramaki", "dL7eDQ9x", "aramaki"),
    100: User(100, "demo", "demo", "demo")
}

# ユーザーチェックに使用する辞書作成
nested_dict = lambda: defaultdict(nested_dict)
user_check = nested_dict()
for i in users.values():
    user_check[i.name]["password"] = i.password
    user_check[i.name]["id"] = i.id


def create_message(from_addr, to_addr, bcc_addrs, subject, body):
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = from_addr
    msg['To'] = to_addr
    msg['Bcc'] = bcc_addrs
    msg['Date'] = formatdate()
    return msg


def send(from_addr, to_addrs, my_pwd, msg):
    smtpobj = smtplib.SMTP('smtp.gmail.com', 587) # gmail
    smtpobj.ehlo()
    smtpobj.starttls()
    smtpobj.ehlo()
    smtpobj.login(from_addr, my_pwd)
    smtpobj.sendmail(from_addr, to_addrs, msg.as_string())
    smtpobj.close()



@app.route('/AccountToroku',methods=["GET", "POST"])
def SendMail_AccountToroku():
  vals = request.json["data"]
  try:
    msg = create_message(mail_address, mail_address, "", "アカウント登録申請", vals[0] + ", " + vals[1])
    send(mail_address, mail_address, mail_password, msg)
    return "0"
  except:
    # 何もしない
    import traceback  
  return "-1"

@login_manager.user_loader
def load_user(user_id):
  return users.get(int(user_id))

# db_uri = "postgresql://postgres:yjrhr1102@localhost:5432/newdb3" #開発用
db_uri = os.environ.get('DATABASE_URL') #本番用
app.config['SQLALCHEMY_DATABASE_URI'] = db_uri 
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
ma.init_app(app)
# q = Queue(connection=conn)
        
@app.route("/favicon.ico")
def favicon():
    return app.send_static_file("favicon.ico")
    


@app.route('/getKaniDateList/<customerId>/<nentuki>')
@login_required
def getKaniDateList(customerId, nentuki):
  y = int(nentuki[0:4])
  m = int(nentuki[4:6])
  
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.customer_id, "
  sql = sql + "     min(c.name1) customer_name1, "
  sql = sql + "     sum(s.price*quantity) sum_price, "
  sql = sql + "     s.deliver_ymd "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "     s.customer_id = c.id "
  sql = sql + "     and s.customer_id = " + customerId + " "
  sql = sql + "     and cast(to_char(deliver_ymd,'yyyy') as integer) = " + str(y) + " "
  sql = sql + "     and cast(to_char(deliver_ymd,'mm') as integer) = " + str(m) + " "
  sql = sql + " group by "
  sql = sql + "     s.customer_id, "
  sql = sql + "     s.deliver_ymd "
  sql = sql + " order by s.deliver_ymd, s.customer_id "
  
  resultset=[]
  data_listA = None
  exist = False

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listA = db.session.execute(text(sql))

    if data_listA is not None:
      for row in data_listA:
        tstr = row["deliver_ymd"].strftime('%Y/%m/%d')
        resultset.append({"deliverYmd":tstr, "customerId":row["customer_id"], "customerName":row["customer_name1"], "sumPrice":row["sum_price"]})

  for d in range(1, 32):
    exist = False
    y = int(nentuki[0:4])
    m = int(nentuki[4:6])
    if isDate(y, m, d):
      deliverymdstr="%04d/%02d/%02d"%(y,m,d)

      for r in resultset:
        if r["deliverYmd"] == deliverymdstr:
          exist = True
          break
      
      if exist == False:
        resultset.append({"deliverYmd":deliverymdstr, "customerId":"", "customerName":"", "sumPrice":""})

  return jsonify({'data': resultset})



@app.route('/getCustomer_Main/<group_kb>/<yuko_muko>/<nen>/<tuki>')
@login_required
def resJson_getCustomer_Main(group_kb, yuko_muko, nen, tuki):
      
      sql = " "
      sql = sql + " SELECT "
      sql = sql + "     c.*, "
      sql = sql + "     k.kakute_ymdt "
      sql = sql + " from "
      sql = sql + "    " + TableWhereTenantId("customer") + " c "
      sql = sql + " left join "
      sql = sql + "    (select * from " + TableWhereTenantId("kakute") + " A where nen = " + nen + " and tuki = " + tuki + ") k "
      sql = sql + " on "
      sql = sql + "     c.id = k.customer_id "
      sql = sql + " where "
      sql = sql + "     c.group_id = " + group_kb + " "
      if yuko_muko == "2":
        sql = sql + "     "
      elif yuko_muko == "1":
        sql = sql + "  and   c.list is not null "
      else:
        sql = sql + "  and   c.list is null "
      sql = sql + "  order by list "
      
      customernentuki = db.session.execute(text(sql))
      customernentuki_schema = CustomerNentukiSchema(many=True)
      return jsonify({'data': customernentuki_schema.dumps(customernentuki, ensure_ascii=False)})


@app.route('/getItem_Daicho/<itemname1>')
@login_required
def resJson_getItem_Daicho(itemname1):
    if itemname1=="すべて":
      items = Item.query.filter(Item.del_flg==0, Item.tenant_id==current_user.tenant_id).all()
    else:
      items = Item.query.filter(Item.del_flg==0,Item.name1==itemname1, Item.tenant_id==current_user.tenant_id).all()

    items_schema = ItemSchema(many=True)
    return jsonify({'data': items_schema.dumps(items, ensure_ascii=False)})

     
@app.route('/getItemGroup_Daicho/')
@login_required
def resJson_getItemGroup_Daicho():
     itemgroup = VItemGroup.query.filter(VItemGroup.tenant_id==current_user.tenant_id).all()
     itemsgroup_schema = VItemGroupSchema(many=True)
     return jsonify({'data': itemsgroup_schema.dumps(itemgroup, ensure_ascii=False)})

     

@app.route('/getVDaichoA_ByCusotmerId/<customerid>')
@login_required
def resJson_getVDaichoA_ByCusotmerId(customerid):
      daicho = VDaichoA.query.filter(VDaichoA.customer_id==customerid, VDaichoA.tenant_id==current_user.tenant_id).all()
      daicho_schema = VDaichoASchema(many=True)
      return jsonify({'data': daicho_schema.dumps(daicho, ensure_ascii=False)})

@app.route('/getVSeikyuA_ByCusotmerIdAndTuki/<customerid>/<nentuki>')
@login_required
def resJson_getVSeikyuA_ByCusotmerId(customerid, nentuki):
      seikyu = VSeikyuA.query.filter(VSeikyuA.customer_id==customerid, VSeikyuA.nen==nentuki[0:4], VSeikyuA.tuki==nentuki[4:6], VSeikyuA.tenant_id==current_user.tenant_id).all()
      seikyu_schema = VSeikyuASchema(many=True)
      return jsonify({'data': seikyu_schema.dumps(seikyu, ensure_ascii=False)})

@app.route('/getSeikyu_ByCusotmerIdAndDate/<customerid>/<deliverYmd>')
@login_required
def resJson_getSeikyu_ByCusotmerIdAndDate(customerid, deliverYmd):
  vDeliverYmd = deliverYmd.replace("-","/")
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.item_id, "
  sql = sql + "     s.price, "
  sql = sql + "     s.quantity, "
  sql = sql + "     i.name1 item_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("item") + " i "
  sql = sql + " where "
  sql = sql + "     s.item_id = i.id "
  sql = sql + "     and s.customer_id = " + customerid + " "
  sql = sql + "     and deliver_ymd = '" + vDeliverYmd + "' "
  # sql = sql + "     and cast(to_char(deliver_ymd,'yyyymmdd') as integer) = " + vDeliverYmd + " "
  sql = sql + " order by s.item_id "
  
  resultset=[]
  data_listA = None

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listA = db.session.execute(text(sql))

    if data_listA is not None:
      for row in data_listA:
        resultset.append({
          "id":row["item_id"], 
          "code":row["item_id"], 
          "tanka":row["price"], 
          "suryo":row["quantity"],
          "name1":row["item_name1"],
          "shokei": int(row["price"]) * int(row["quantity"]),
        })
  return jsonify({'data': resultset})

@app.route('/OutputExcelNouhinsho/<customerid>/<deliverYmd>')
@login_required
def resExcelFile_OutputExcelNouhinsho(customerid, deliverYmd):
  resultset=[]
  
  vDeliverYmd = deliverYmd.replace("-","/")
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.item_id, "
  sql = sql + "     s.price, "
  sql = sql + "     s.quantity, "
  sql = sql + "     i.name1 item_name1, "
  sql = sql + "     c.name1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("item") + " i, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "     s.item_id = i.id "
  sql = sql + "     and s.customer_id = c.id "
  sql = sql + "     and s.customer_id = " + customerid + " "
  sql = sql + "     and deliver_ymd = '" + vDeliverYmd + "' "
  # sql = sql + "     and cast(to_char(deliver_ymd,'yyyymmdd') as integer) = " + vDeliverYmd + " "
  sql = sql + " order by s.item_id "
  
  resultset=[]
  data_listA = None

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listA = db.session.execute(text(sql))

    if data_listA is not None:
      for row in data_listA:
        resultset.append({
          "item_id":row["item_id"], 
          "price":row["price"], 
          "quantity":row["quantity"],
          "item_name1":row["item_name1"],
          "customer_name1":row["customer_name1"],
        })
  
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + customerid + "_" + timestampStr + "_" + current_user.name + "_" + current_user.tenant_id
  
  wb = openpyxl.load_workbook('ExcelTemplate/hoiku/納品書_保育園.xlsx')

  if len(resultset) > 0:
    sheet = wb['納品書']
    # cell = sheet['A2']
    sheet['A2'] = resultset[0]["customer_name1"] + "　" + "御中"
    dd = vDeliverYmd.split("/")
    sheet['P2'] = dd[0] + "年 " + dd[1] + "月 " + dd[2] + "日"

    idx = 1
    for r in resultset:
      sheet['A' + str(9 + idx)] = idx
      sheet['C' + str(9 + idx)] = r["item_name1"]
      sheet['L' + str(9 + idx)] = r["quantity"]
      sheet['N' + str(9 + idx)] = r["price"]

      idx += 1

  wb.save('tmp/' + filename + '.xlsx')

  return send_file('tmp/' + filename + '.xlsx', as_attachment=True, mimetype=XLSX_MIMETYPE, attachment_filename = filename + '.xlsx')


@app.route('/OutputExcelOrderSlip/<orderYmd>/<hopeYmd>/<sendStamp>')
@login_required
def resExcelFile_OutputExcelOrderSlip(orderYmd, hopeYmd, sendStamp):
  resultset=[]
  vOrderYmd = orderYmd # .replace("-","/")
  vHopeYmd = hopeYmd # .replace("-","/")
  vSendStamp = sendStamp # .replace("-","/")

  sql = " "
  sql = sql + " select "
  sql = sql + "     oi.item_code item_id, "
  sql = sql + "     oi.item_siire price, "
  sql = sql + "     oi.quantity quantity, "
  sql = sql + "     oi.item_name1 item_name1, "
  sql = sql + "     'さとう牛乳販売店' customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("order_item") + " oi "
  sql = sql + " where "
  sql = sql + "         oi.order_ymd = '" + vOrderYmd + "' "
  sql = sql + "     and oi.hope_ymd = '" + vHopeYmd + "' "
  sql = sql + "     and oi.send_stamp = '" + vSendStamp + "' "
  sql = sql + " order by cast(oi.item_code as integer) "
  
  resultset=[]
  data_listA = None

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listA = db.session.execute(text(sql))

    if data_listA is not None:
      for row in data_listA:
        resultset.append({
          "item_id":row["item_id"], 
          "price":row["price"], 
          "quantity":row["quantity"],
          "item_name1":row["item_name1"],
          "customer_name1":row["customer_name1"],
        })
  
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + timestampStr + "_" + current_user.name + "_" + current_user.tenant_id
  
  wb = openpyxl.load_workbook('ExcelTemplate/order/納品書_日次.xlsx')

  if len(resultset) > 0:
    sheet = wb['納品書']
    # cell = sheet['A2']
    sheet['A2'] = resultset[0]["customer_name1"] + "　" + "御中"
    dd = vHopeYmd.split("-")
    sheet['P2'] = dd[0] + "年 " + dd[1] + "月 " + dd[2] + "日"

    idx = 1
    for r in resultset:
      sheet['A' + str(9 + idx)] = idx
      sheet['C' + str(9 + idx)] = r["item_id"]
      sheet['E' + str(9 + idx)] = r["item_name1"]
      sheet['Q' + str(9 + idx)] = r["quantity"]
      # sheet['N' + str(9 + idx)] = r["price"]

      idx += 1

  wb.save('tmp/' + filename + '.xlsx')

  return send_file('tmp/' + filename + '.xlsx', as_attachment=True, mimetype=XLSX_MIMETYPE, attachment_filename = filename + '.xlsx')

def SeikyuExcelSqlA(nentuki, customerid):
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.deliver_ymd, "
  sql = sql + "     s.item_id, "
  sql = sql + "     s.price, "
  sql = sql + "     s.quantity, "
  sql = sql + "     i.name1 item_name1, "
  sql = sql + "     c.name1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("item") + " i, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "     s.item_id = i.id "
  sql = sql + "     and s.customer_id = c.id "
  sql = sql + "     and s.customer_id = " + str(customerid) + " "
  sql = sql + "     and cast(to_char(s.deliver_ymd,'yyyymm') as integer) = " + nentuki + " "
  sql = sql + " order by c.name1, s.deliver_ymd, s.item_id "
  return sql


def SeikyuExcelSqlD(nentuki, customerid):
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.deliver_ymd, "
  sql = sql + "     s.item_id, "
  sql = sql + "     s.price, "
  sql = sql + "     s.quantity, "
  sql = sql + "     i.name1 item_name1, "
  sql = sql + "     c.name1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("item") + " i, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "     s.item_id = i.id "
  sql = sql + "     and s.customer_id = c.id "
  sql = sql + "     and s.customer_id = " + str(customerid) + " "
  sql = sql + "     and cast(to_char(s.deliver_ymd,'yyyymm') as integer) = " + nentuki + " "
  sql = sql + " order by c.name1, s.item_id, s.deliver_ymd "
  return sql

def SeikyuExcelSqlOrderDetail(dateJoken):
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.order_ymd deliver_ymd, "
  sql = sql + "     s.item_id, "
  sql = sql + "     s.item_siire price, "
  sql = sql + "     s.quantity, "
  sql = sql + "     s.item_name1 item_name1, "
  sql = sql + "     c.param_val1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("order_item") + " s, "
  sql = sql + "     " + TableWhereTenantId("mst_setting") + " c "
  sql = sql + " where "
  sql = sql + "         c.param_id = 'TENPO_SEIKYUSHO' "
  sql = sql + "     and c.param_no = 1 "
  sql = sql + "     and to_char(s.order_ymd,'yyyy-mm') = '" + dateJoken + "' "
  sql = sql + " order by s.item_id, s.order_ymd "
  return sql

def SeikyuExcelSqlB(nentuki, customerid):
  sql = " "
  sql = sql + " select "
  sql = sql + "     s.deliver_ymd, "
  sql = sql + "     sum(s.price * s.quantity) zeinuki, "
  sql = sql + "     s.customer_id, "
  sql = sql + "     c.name1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "         s.customer_id = c.id "
  sql = sql + "     and s.customer_id = " + str(customerid) + " "
  sql = sql + "     and cast(to_char(s.deliver_ymd,'yyyymm') as integer) = " + nentuki + " "
  sql = sql + " group by "
  sql = sql + "     s.deliver_ymd, "
  sql = sql + "     s.customer_id, "
  sql = sql + "     c.name1 "
  sql = sql + " order by s.customer_id, s.deliver_ymd "
  return sql
  

def SeikyuExcelSqlC(nentuki):
  sql = " "
  sql = sql + " select "
  sql = sql + "     sum(s.price * s.quantity) zeinuki, "
  sql = sql + "     s.customer_id, "
  sql = sql + "     c.name1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("seikyu") + " s, "
  sql = sql + "     " + TableWhereTenantId("customer") + " c "
  sql = sql + " where "
  sql = sql + "         s.customer_id = c.id "
  sql = sql + "     and cast(to_char(s.deliver_ymd,'yyyymm') as integer) = " + nentuki + " "
  sql = sql + " group by "
  sql = sql + "     s.customer_id, "
  sql = sql + "     c.name1 "
  sql = sql + " order by s.customer_id "
  return sql
  

def SeikyuExcelSqlOrder(dateJoken):
  # vfrom = (dateFrom.replace("-",""))
  # vto = (dateTo.replace("-",""))
  sql = " "
  sql = sql + " select "
  sql = sql + "     sum(s.item_siire * s.quantity) zeinuki, "
  sql = sql + "     c.param_no customer_id, "
  sql = sql + "     c.param_val1 customer_name1 "
  sql = sql + " from "
  sql = sql + "     " + TableWhereTenantId("order_item") + " s, "
  sql = sql + "     " + TableWhereTenantId("mst_setting") + " c "
  sql = sql + " where "
  sql = sql + "         c.param_id = 'TENPO_SEIKYUSHO' "
  sql = sql + "     and c.param_no = 1 "
  sql = sql + "     and to_char(s.order_ymd,'yyyy-mm') = '" + dateJoken + "' "
  sql = sql + " group by "
  sql = sql + "     c.param_no, "
  sql = sql + "     c.param_val1 "
  sql = sql + " order by 1 "
  return sql
  

@app.route('/OutputExcelSeikyusho/<nentuki>')
@login_required
def resExcelFile_OutputExcelSeikyusho(nentuki):
  
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + nentuki + "_" + timestampStr + "_" + current_user.name + "_" + current_user.tenant_id
  
  wb = openpyxl.load_workbook('ExcelTemplate/hoiku/請求書_指定.xlsx')

  resultsetC=[]
  data_listC = None
  sql = SeikyuExcelSqlC(nentuki)

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listC = db.session.execute(text(sql))

    if data_listC is not None:
      for row in data_listC:
        resultsetC.append({
          "zeinuki":row["zeinuki"], "customer_id":row["customer_id"], "customer_name1":row["customer_name1"],
        })

  ccnt = 0
  if len(resultsetC) > 0:
    for c in resultsetC:
      
      sheet = wb.worksheets[ccnt]
      sheet.title = c["customer_name1"]
      sheet['AZ36'] = "09020721"

      ccnt = ccnt + 1
      sheet['I' + str(3 + (65 * 1))] = "　" + c["customer_name1"] + "　様"
      # page_break = Break(id=10) # create Break obj
      # sheet.page_breaks[0].append(page_break) # insert page break

      
      resultsetA=[]
      data_listA = None
      sql = SeikyuExcelSqlA(nentuki, c["customer_id"])

      if db.session.execute(text(sql)).fetchone() is not None:
        data_listA = db.session.execute(text(sql))

        if data_listA is not None:
          for row in data_listA:
            resultsetA.append({
              "deliver_ymd":row["deliver_ymd"], "item_id":row["item_id"], 
              "price":row["price"], "quantity":row["quantity"],
              "item_name1":row["item_name1"], "customer_name1":row["customer_name1"],
            })


      prevDate = ""
      prevDateCellSt = ""
      prevDateCellEn = ""
      nikkei = 0
      idx = 7 + 65 #(65 * ccnt)
      for r in resultsetA:
        if prevDate != r["deliver_ymd"].strftime('%m/%d') :
          sheet['C' + str(idx)] = r["deliver_ymd"].strftime('%m/%d') 

          if prevDateCellSt != "" and prevDateCellEn != "" and int(prevDateCellSt.replace("BE","")) <= int(prevDateCellEn.replace("BM","")) :
            sheet.merge_cells(prevDateCellSt + ":" + prevDateCellEn)
            sheet[prevDateCellSt] = nikkei + math.floor(nikkei*0.08) #4567
            sheet[prevDateCellSt].alignment = openpyxl.styles.Alignment(wrapText=True, vertical = 'center')
            sheet.merge_cells(prevDateCellSt.replace("BE","C") + ":" + prevDateCellEn.replace("BM","F"))
            sheet.merge_cells(prevDateCellSt.replace("BE","AY") + ":" + prevDateCellEn.replace("BM","BD"))
            sheet[prevDateCellSt.replace("BE","AY")] = math.floor(nikkei*0.08) #消費税
            nikkei = 0
            sheet[prevDateCellSt.replace("BE","C")].alignment = openpyxl.styles.Alignment(vertical = 'top', horizontal="center")
            prevDateCellEn = ""

          prevDateCellSt = 'BE' + str(idx)
        
        prevDateCellEn = 'BM' + str(idx)

        sheet['G' + str(idx)] = r["item_name1"]
        sheet['AC' + str(idx)] = r["quantity"]
        sheet['AG' + str(idx)] = r["price"]
        sheet['AM' + str(idx)] = int(r["price"]) * int(r["quantity"])
        nikkei = nikkei + int(r["price"]) * int(r["quantity"])
        sheet['AU' + str(idx)] = "8%"

        prevDate = r["deliver_ymd"].strftime('%m/%d')

        idx += 1

      sheet.merge_cells(prevDateCellSt + ":" + prevDateCellEn)
      sheet[prevDateCellSt] = nikkei + math.floor(nikkei*0.08) #4567
      sheet[prevDateCellSt].alignment = openpyxl.styles.Alignment(wrapText=True, vertical = 'center')
      sheet.merge_cells(prevDateCellSt.replace("BE","C") + ":" + prevDateCellEn.replace("BM","F"))
      sheet.merge_cells(prevDateCellSt.replace("BE","AY") + ":" + prevDateCellEn.replace("BM","BD"))
      sheet[prevDateCellSt.replace("BE","AY")] = math.floor(nikkei*0.08) #消費税
      nikkei = 0
      sheet[prevDateCellSt.replace("BE","C")].alignment = openpyxl.styles.Alignment(vertical = 'top', horizontal="center")
      prevDateCellEn = ""

      resultsetB=[]
      data_listB = None
      sql = SeikyuExcelSqlB(nentuki, c["customer_id"])

      if db.session.execute(text(sql)).fetchone() is not None:
        data_listB = db.session.execute(text(sql))

        if data_listB is not None:
          for row in data_listB:
            resultsetB.append({
              "deliver_ymd":row["deliver_ymd"], "zeinuki":row["zeinuki"], 
              "customer_id":row["customer_id"], "customer_name1":row["customer_name1"],
            })
              
            if len(resultsetB) > 0:
              gokei = 0
              idx = 4
              for r in resultsetB:
                sheet['BQ' + str(idx)] = r["deliver_ymd"]
                sheet['BR' + str(idx)] = r["customer_name1"]
                sheet['BS' + str(idx)] = int(r["zeinuki"])
                sheet['BT' + str(idx)] = math.floor(int(r["zeinuki"])*0.08)
                sheet['BU' + str(idx)] = int(r["zeinuki"]) + math.floor(int(r["zeinuki"])*0.08)
                idx += 1
                gokei = gokei + (int(r["zeinuki"]) + math.floor(int(r["zeinuki"])*0.08))

              kingakuarray = list('￥' + str(gokei))
              kingakuarray.reverse()
              cellkey = ["AY10", "AV10", "AS10", "AP10", "AM10", "AJ10", "AG10"]
              for k in range(7):
                if k < len(kingakuarray):
                  sheet[cellkey[k]] = kingakuarray[k]

  wb.save('tmp/' + filename + '.xlsx')

  return send_file('tmp/' + filename + '.xlsx', as_attachment=True, mimetype=XLSX_MIMETYPE, attachment_filename = filename + '.xlsx')


@app.route('/OutputExcelSeikyushoB/<nentuki>')
@login_required
def resExcelFile_OutputExcelSeikyushoB(nentuki):
  
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + nentuki + "_" + timestampStr + "_" + current_user.name + "_" + current_user.tenant_id
  
  wb = openpyxl.load_workbook('ExcelTemplate/hoiku/請求書_指定B.xlsx')

  resultsetC=[]
  data_listC = None
  sql = SeikyuExcelSqlC(nentuki)

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listC = db.session.execute(text(sql))

    if data_listC is not None:
      for row in data_listC:
        resultsetC.append({
          "zeinuki":row["zeinuki"], "customer_id":row["customer_id"], "customer_name1":row["customer_name1"],
        })

  ccnt = 0
  if len(resultsetC) > 0:
    for c in resultsetC:
      
      # sheet = wb.worksheets[ccnt]
      sheet = wb.copy_worksheet(wb['Sheet1'])
      sheet.title = c["customer_name1"]

      ccnt = ccnt + 1
      sheet['A1'] = "　" + c["customer_name1"] + "　様"
      sheet['F4'] = c["zeinuki"] 
      sheet['A4'] = nentuki[0:4] + " 年 " + nentuki[4:6] + " 月 分"
      
      resultsetA=[]
      data_listA = None
      sql = SeikyuExcelSqlD(nentuki, c["customer_id"])

      if db.session.execute(text(sql)).fetchone() is not None:
        data_listA = db.session.execute(text(sql))

        if data_listA is not None:
          for row in data_listA:
            resultsetA.append({
              "deliver_ymd":row["deliver_ymd"], "item_id":row["item_id"], 
              "price":row["price"], "quantity":row["quantity"],
              "item_name1":row["item_name1"], "customer_name1":row["customer_name1"],
            })

      itemColumnId = ["B","D","F","H","J","L","N","P"]
      nikkei = 0
      gyoNum = 0
      idx = 1
      itemIndex = -1
      prevItemId = 0
      for r in resultsetA:
        if prevItemId != r["item_id"]:
          itemIndex += 1
          sheet[itemColumnId[itemIndex] + "7"] = r["item_name1"]
          sheet[itemColumnId[itemIndex] + "8"] = r["price"]

        gyoNum = int(r["deliver_ymd"].strftime('%d')) + 10
        sheet[itemColumnId[itemIndex] + str(gyoNum)] = r["quantity"]
        prevItemId = r["item_id"]
        idx += 1

  wb.remove(wb['Sheet1'])
  wb.save('tmp/' + filename + '.xlsx')

  return send_file('tmp/' + filename + '.xlsx', as_attachment=True, mimetype=XLSX_MIMETYPE, attachment_filename = filename + '.xlsx')


@app.route('/getCustomer_ById/<customerid>')
@login_required
def resJson_getCustomer_ById(customerid):
      customer = Customer.query.filter(Customer.id==customerid, Customer.tenant_id==current_user.tenant_id).all()
      customer_schema = CustomerSchema(many=True)
      return jsonify({'data': customer_schema.dumps(customer, ensure_ascii=False)})


@app.route('/getItem_ById/<itemid>')
@login_required
def resJson_getItem_ById(itemid):
      item = Item.query.filter(Item.id==itemid, Item.tenant_id==current_user.tenant_id).all()
      item_schema = ItemSchema(many=True)
      return jsonify({'data': item_schema.dumps(item, ensure_ascii=False)})

@app.route('/getDaicho_ByItemId/<itemid>')
@login_required
def resJson_getDaicho_ByItemId(itemid):
      daicho = VDaichoA.query.filter(VDaichoA.item_id==itemid, VDaichoA.tenant_id==current_user.tenant_id).all()
      daicho_schema = VDaichoASchema(many=True)
      return jsonify({'data': daicho_schema.dumps(daicho, ensure_ascii=False)})

@app.route('/getSeikyuNengetuShukei_Main')
@login_required
def resJson_getSeikyuNengetuShukei_Main():
      seikyu = VSeikyuC.query.filter(VSeikyuC.tenant_id==current_user.tenant_id).all()
      seikyu_schema = VSeikyuCSchema(many=True)
      return jsonify({'data': seikyu_schema.dumps(seikyu, ensure_ascii=False)})


@app.route('/getSeikyuNengetuCustomer_Main/<nen>/<tuki>/<groupkb>')
@login_required
def resJson_getSeikyuNengetuCustomer_Main(nen, tuki, groupkb):
      seikyu = VSeikyuB.query.filter(VSeikyuB.nen==nen, VSeikyuB.tuki==tuki, VSeikyuB.group_id==groupkb, VSeikyuB.tenant_id==current_user.tenant_id).all()
      seikyu_schema = VSeikyuBSchema(many=True)
      return jsonify({'data': seikyu_schema.dumps(seikyu, ensure_ascii=False)})



@app.route('/createSeikyu/<customerid>/<nentuki>/<sakujonomi>')
@login_required
def dbUpdate_insSeikyu(customerid, nentuki, sakujonomi):
  y = int(nentuki[0:4])
  m = int(nentuki[4:6])
  
  sql = " "
  sql = sql + " delete from seikyu "
  sql = sql + " where tenant_id = '" + current_user.tenant_id + "' "
  if customerid != '-1' :
    sql = sql + "     and customer_id = " + customerid + " "
  
  sql = sql + "     and cast(to_char(deliver_ymd,'yyyy') as integer) = " + str(y) + " "
  sql = sql + "     and cast(to_char(deliver_ymd,'mm') as integer) = " + str(m) + " "
  db.session.execute(text(sql))
  
  if sakujonomi == 'true' :
    db.session.commit()
    return "1"
  
  blAri = False
  for d in range(1,32):
    if isDate(y, m, d):
      deliverymdstr="%04d/%02d/%02d"%(y,m,d)
      deliverymd=datetime.datetime.strptime(deliverymdstr,"%Y/%m/%d")
      
      sql = " "
      sql = sql + " SELECT "
      sql = sql + "     d.customer_id, "
      sql = sql + "     to_date('" + deliverymdstr + "','yyyy/mm/dd') deliver_ymd, "
      sql = sql + "     d.item_id, "
      sql = sql + "     i.tanka, "
      sql = sql + "     null price_sub, "
      sql = sql + "     d.quantity, "
      sql = sql + "     'dummy' user_id, "
      sql = sql + "     CURRENT_TIMESTAMP "
      sql = sql + " from "
      sql = sql + "    " + TableWhereTenantId("daicho") + " d "
      sql = sql + " inner join "
      sql = sql + "    " + TableWhereTenantId("customer") + " c "
      sql = sql + " on "
      sql = sql + "     d.customer_id =  c.id "
      sql = sql + " inner join "
      sql = sql + "    " + TableWhereTenantId("item") + " i "
      sql = sql + " on "
      sql = sql + "     d.item_id =  i.id "
      sql = sql + " where "
      if customerid != '-1' :
        sql = sql + "     d.customer_id = " + customerid + " and "
      sql = sql + "     d.youbi = " + str(deliverymd.weekday()+1) + " and "
      sql = sql + "     c.list is not null and "
      sql = sql + "     c.list <> 0 "
      # print(sql)
      
      # print(db.session.execute(text(sql)).fetchone())
      
      if db.session.execute(text(sql)).fetchone() is not None:
        # print(db.session.execute(text(sql)).fetchone())
        blAri = True
        data_list = db.session.execute(text(sql))
        seikyus = [{'customer_id':d[0], 'deliver_ymd': d[1], 'item_id': d[2],
                  'price': d[3], 'price_sub': d[4], 'quantity': d[5], 'user_id': current_user.name, 'ymdt': d[7], 'tenant_id': current_user.tenant_id} for d in data_list]
                  
        db.session.execute(Seikyu.__table__.insert(), seikyus)
        db.session.commit()
  
  if blAri :
    return str(customerid)
  else :
    return "-1"
  
def isDate(year,month,day):
    try:
        newDataStr="%04d/%02d/%02d"%(year,month,day)
        newDate=datetime.datetime.strptime(newDataStr,"%Y/%m/%d")
        return True
    except ValueError:
        return False


def TableWhereTenantId(table_nm):
  return " (select * from " + table_nm + " where tenant_id = '" + current_user.tenant_id + "') "



@app.route('/printSeikyu/<customerid>/<customeridB>/<nentuki>/<randnum>')
@login_required
def resPdf_printSeikyu(customerid, customeridB, nentuki, randnum):
  # 
  sql = ""
  sql = sql + "  SELECT to_char(seikyu.deliver_ymd,'yyyy')        nen,                                                                                  " 
  sql = sql + "         to_char(seikyu.deliver_ymd,'mm')         tuki,                                                                                   " 
  sql = sql + "         seikyu.customer_id,                                                                                                             " 
  sql = sql + "         seikyu.deliver_ymd,                                                                                                             " 
  sql = sql + "         seikyu.item_id,                                                                                                                 " 
  sql = sql + "         seikyu.price,                                                                                                                   " 
  sql = sql + "         seikyu.quantity,                                                                                                                " 
  sql = sql + "         item.code                               item_code,                                                                              " 
  sql = sql + "         item.name1                              item_name1,                                                                             " 
  sql = sql + "         item.name2                              item_name2,                                                                             " 
  sql = sql + "         customer.name1                          customer_name1,                                                                         " 
  sql = sql + "         customer.name2                          customer_name2,                                                                         " 
  sql = sql + "         customer.list,                                                                                                                  " 
  sql = sql + "         customer.group_id,                                                                                                              " 
  sql = sql + "         to_char(seikyu.deliver_ymd,'yyyy') || to_char(seikyu.deliver_ymd,'mm') || lpad(seikyu.customer_id::text,6,0::text) SEIKYU_KEY,  " 
  sql = sql + "         customer.harai_kb ,                                                                                                             " 
  sql = sql + "         customer.biko2 zei_kb,                                                                                                           " 
  sql = sql + "         customer.biko1 tanto                                                                                                          " 
  sql = sql + "  FROM   " + TableWhereTenantId("seikyu") + " seikyu                                                                                     " 
  sql = sql + "  inner join " + TableWhereTenantId("item") + " item                                                                                     " 
  sql = sql + "  on                                                                                                                                     " 
  sql = sql + "      seikyu.item_id = item.id                                                                                                           " 
  sql = sql + "  inner join " + TableWhereTenantId("customer") + " customer                                                                             " 
  sql = sql + "  on                                                                                                                                     " 
  sql = sql + "      seikyu.customer_id = customer.id                                                                                                   " 
  sql = sql + "  left outer join " + "(select distinct customer_id, item_id from " + TableWhereTenantId("daicho") + " daicho) daicho                    " 
  sql = sql + "  on                                                                                                                                     " 
  sql = sql + "      seikyu.customer_id = daicho.customer_id                                                                                                   " 
  sql = sql + "      and seikyu.item_id = daicho.item_id                                                                                                   " 
  sql = sql + "  where                                                                                                                                  " 
  sql = sql + "       to_char(seikyu.deliver_ymd,'yyyy') = '" + nentuki[0:4] + "' and                                                                   " 
  sql = sql + "       to_char(seikyu.deliver_ymd,'mm') = '" + nentuki[4:6] + "' and                                                                     " 
  sql = sql + "       seikyu.customer_id = V_CUSTOMER_ID_V and                                                                                       " 
  sql = sql + "       list IS NOT NULL                                                                                                                  " 
  sql = sql + "  ORDER  BY to_char(seikyu.deliver_ymd,'yyyy'),                                                                                          " 
  sql = sql + "            to_char(seikyu.deliver_ymd,'mm'),                                                                                            " 
  sql = sql + "            customer.list,                                                                                                               " 
  sql = sql + "            seikyu.customer_id,                                                                                                          " 
  sql = sql + "            daicho.item_id asc nulls last,                                                                                                              " 
  sql = sql + "            seikyu.item_id,                                                                                                              " 
  sql = sql + "            seikyu.deliver_ymd;                                                                                                          " 

  sqlA = sql.replace("V_CUSTOMER_ID_V",customerid)
  sqlB = sql.replace("V_CUSTOMER_ID_V",customeridB)
  # sql = " select * from v_seikyu_b where nen = '2021' and tuki = '02' and customer_id = " + customerid

  param_list = MstSetting.query.filter(MstSetting.tenant_id==current_user.tenant_id).all()

  if db.session.execute(text(sqlA)).fetchone() is not None:
    data_listA = db.session.execute(text(sqlA))

    if db.session.execute(text(sqlB)).fetchone() is not None:
      data_listB = db.session.execute(text(sqlB))
    else:
      data_listB = None
    
    timestamp = datetime.datetime.now()
    timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
    filename = "file_" + customerid + "_" + customeridB + "_" + timestampStr + "_" + current_user.name
    make(filename, data_listA, data_listB, param_list)

    response = make_response()
    response.data = open("tmp/" + filename + ".pdf", "rb").read()
    response.headers['Content-Disposition'] = "attachment; filename=unicode.pdf"
    response.mimetype = 'application/pdf'
    return filename + ".pdf"
  else:
    return "-1"


@app.route('/pdfMergeSeikyusho',methods=["GET", "POST"])
@login_required
def print_pdfMergeSeikyusho():
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  vals = request.json["data"]
  merger = PyPDF2.PdfFileMerger()

  idx = 0
  tryCnt = 0
  while True :
    try:
      merger.append("tmp/" + vals[0] + "")
      print("成功：" + vals.pop(0))
    except:
      print("失敗：" + vals[0])
    finally:
      tryCnt = tryCnt + 1

    if len(vals)==0 :
      break
    if tryCnt > 9999 :
      break

  merger.write("tmp/" + timestampStr + ".pdf")
  merger.close()
  
  return send_file("tmp/" + timestampStr + ".pdf", as_attachment=True)

@app.route('/getMstSetting_Main/<param_id>')
@login_required
def resJson_getMstSetting_Main(param_id):
  setting = MstSetting.query.filter(MstSetting.param_id==param_id, 
  MstSetting.tenant_id==current_user.tenant_id).order_by(asc(MstSetting.param_id),asc(MstSetting.param_no)).all() #変更
  setting_schema = MstSettingSchema(many=True)
  return jsonify({'data': setting_schema.dumps(setting, ensure_ascii=False)})

@app.route('/isKanriUser')
@login_required
def resJson_isKanriUser():
  setting = MstSetting.query.filter(MstSetting.param_id=="KANRI_USER", MstSetting.tenant_id==current_user.tenant_id, MstSetting.param_val1==current_user.name).all() 
  setting_schema = MstSettingSchema(many=True)
  return jsonify({'data': setting_schema.dumps(setting, ensure_ascii=False)})

@app.route('/getMstSetting_Full')
@login_required
def resJson_getMstSetting_Full():
  setting = MstSetting.query.distinct(MstSetting.param_id, MstSetting.param_nm).filter(MstSetting.tenant_id==current_user.tenant_id).all()
  setting_schema = MstSettingSchema(many=True)
  return jsonify({'data': setting_schema.dumps(setting, ensure_ascii=False)})






@app.route('/updateSetteiText/<params>')
@login_required
def dbUpdate_updateSetteiText(params):
  vals = params.split(",")
  # param_id, param_nm, param_no, param_val1, param_val2, colIndex, val
  MstSetting.query.filter( \
    MstSetting.param_id==vals[0], \
    MstSetting.param_no==vals[2], \
    MstSetting.tenant_id==current_user.tenant_id).delete()

  mstsetting = MstSetting()
  mstsetting.param_id = vals[0]
  mstsetting.param_nm = vals[1]
  mstsetting.param_no = vals[2]
  mstsetting.param_val1 = null2blank(vals[6]) if int(vals[5])==1 else null2blank(vals[3])  #param_val1 #"OK" if n == 10 else "NG"
  mstsetting.param_val2 = null2blank(vals[6]) if int(vals[5])==2 else null2blank(vals[4])  #param_val2
  mstsetting.param_val3 = ""
  mstsetting.tenant_id = current_user.tenant_id
  db.session.add(mstsetting)

  # データを確定
  db.session.commit()
  return "1"




@app.route('/updateKakute/<nen>/<tuki>/<customerid>')
@login_required
def dbUpdate_updateKakute(nen, tuki, customerid):
  kakute = Kakute.query.filter( \
    Kakute.nen == nen, \
    Kakute.tuki == tuki, \
    Kakute.customer_id == customerid, \
    Kakute.tenant_id==current_user.tenant_id).all()
  
  delOnly=False
  if len(kakute)==1:
    if kakute[0].kakute_ymdt != None:
      delOnly=True

  Kakute.query.filter( \
    Kakute.nen == nen, \
    Kakute.tuki == tuki, \
    Kakute.customer_id == customerid, \
    Kakute.tenant_id==current_user.tenant_id).delete()
  
  if delOnly==False:
    kakute = Kakute()
    kakute.nen = nen
    kakute.tuki = tuki
    kakute.customer_id = customerid
    kakute.tenant_id = current_user.tenant_id
    kakute.kakute_ymdt = datetime.datetime.now()
    db.session.add(kakute)
# 
  # # データを確定
  db.session.commit()
  return customerid



def null2blank(val):
  if val == "null":
    return ""
  else:
    return val


@app.route('/getDaichoCustomer_SeikyuSub')
@login_required
def resJson_getDaichoCustomer_SeikyuSub():
  customer = Customer.query.filter(Customer.list!=None, Customer.tenant_id==current_user.tenant_id).all()
  customer_schema = CustomerSchema(many=True)
  return jsonify({'data': customer_schema.dumps(customer, ensure_ascii=False)})

@app.route('/updAddDaicho/<param>')
@login_required
def dbUpdate_updAddDaicho(param):
  vals = param.split(",")
  # print(vals)
  Daicho.query.filter(Daicho.quantity==0, Daicho.tenant_id==current_user.tenant_id).delete()
  for youbi in range(2, 9):
    Daicho.query.filter(Daicho.customer_id==vals[0], Daicho.item_id==vals[1], Daicho.youbi==(youbi-1), Daicho.tenant_id==current_user.tenant_id).delete()
    if vals[youbi].isdecimal():
      if int(vals[youbi]) != 0 :
        daicho = Daicho()
        daicho.customer_id = vals[0]
        daicho.item_id = vals[1]
        daicho.youbi = (youbi-1)
        daicho.quantity = vals[youbi]
        daicho.tenant_id = current_user.tenant_id
        db.session.add(daicho)
    db.session.commit()
  return param

@app.route('/updTakuhaijun',methods=["GET", "POST"])
@login_required
def dbUpdate_updTakuhaijun():
  vals = request.json["data"]
  for id_list in vals:
    customer = Customer.query.filter(Customer.id==id_list[0], Customer.tenant_id==current_user.tenant_id).first()
    customer.list = id_list[1]
    if str(id_list[1]) == "None":
      customer.address3 = None
    else:
      customer.address3 = str(id_list[1])
  db.session.commit()
  return "1"


@app.route('/updateSeikyuQuantity/<customerid>/<itemid>/<deliverymd>/<quantity>/<price>/<pricesub>')
@login_required
def dbUpdate_updSeikyuQuantity(customerid, itemid, deliverymd, quantity, price, pricesub):
  
  Seikyu.query.filter(Seikyu.customer_id==customerid, Seikyu.item_id==itemid, Seikyu.deliver_ymd==deliverymd, Seikyu.tenant_id==current_user.tenant_id).delete()

  if quantity == "b":
    # データを確定
    db.session.commit()
    return "1"

  tstr = deliverymd #'2012-12-29 13:49:37'
  tdatetime = datetime.datetime.strptime(tstr, '%Y-%m-%d')
  tdate = datetime.date(tdatetime.year, tdatetime.month, tdatetime.day)

  d = Daicho.query.filter(Daicho.customer_id==customerid, Daicho.item_id==itemid, Daicho.youbi==str(tdate.weekday()+1), Daicho.tenant_id==current_user.tenant_id).all()
  
  if int(quantity) != 0  or len(d) != 0:
    seikyu = Seikyu()
    seikyu.customer_id = customerid
    seikyu.item_id = itemid
    seikyu.deliver_ymd = deliverymd
    seikyu.price = price
    seikyu.price_sub = price if price == "null" else None
    seikyu.quantity = int(quantity)
    seikyu.user_id = current_user.name
    seikyu.ymdt = datetime.datetime.now()
    seikyu.tenant_id = current_user.tenant_id
    db.session.add(seikyu)

  # データを確定
  db.session.commit()
  return "1"





@app.route('/UpdateItem/<param>')
@login_required
def dbUpdate_UpdateItem(param):
  vals = param.split(DELIMIT)
  itemid = int(vals[0])
  if itemid == 0 :  #新規登録
    item = Item()
    item.code = vals[1]
    item.name1 = vals[2]
    item.name2 = vals[3]
    item.tanka = int(vals[4])
    item.orosine = vals[5]
    item.zei_kb = int(vals[6])
    item.del_flg = int(vals[7])
    item.tenant_id = current_user.tenant_id
    db.session.add(item)
  else:
    item = Item.query.filter(Item.id==itemid, Item.tenant_id==current_user.tenant_id).first()
    item.code = vals[1]
    item.name1 = vals[2]
    item.name2 = vals[3]
    item.tanka = int(vals[4])
    item.orosine = vals[5]
    item.zei_kb = int(vals[6])
    item.del_flg = int(vals[7])
    item.tenant_id = current_user.tenant_id

  # データを確定
  db.session.commit()
  return param



@app.route('/getCsvData/<viewnm>/<nentuki>/<groupkb>/<tanto>')
@login_required
def resJson_getCsvData(viewnm, nentuki, groupkb, tanto):

  sqlwhere=" tenant_id = '" + current_user.tenant_id + "' "
  if viewnm == "v_csv_uriage_tantobetu":
    sqlwhere = sqlwhere + " and nen = '" + nentuki[0:4] + "' and tuki = '" + nentuki[4:6] + "' and group_id = " + groupkb + " and tanto_id = '" + tanto + "' " 
  elif viewnm == "v_csv_uriage_groupbetu":
    sqlwhere = sqlwhere + " and nen = '" + nentuki[0:4] + "' and tuki = '" + nentuki[4:6] + "' and group_id = " + groupkb + " and tanto_id = '" + tanto + "' " 
  elif viewnm == "v_csv_uriage_kokyakubetu":
    sqlwhere = sqlwhere + " and nen = '" + nentuki[0:4] + "' and tuki = '" + nentuki[4:6] + "' and group_id = " + groupkb + " and tanto_id = '" + tanto + "' " 
  elif viewnm == "v_csv_hikiotosi":
    sqlwhere = " nen = '" + nentuki[0:4] + "' and tuki = '" + nentuki[4:6] + "' order by nen, tuki, harai_kb, name2 COLLATE \"ja_JP.utf8\", list " 
  elif viewnm == "v_csv_takuhai":
    sqlwhere = sqlwhere + " and group_id = " + groupkb + " and tanto_id = '" + tanto + "' " 
  elif viewnm == "v_csv_daily":
    sqlwhere = " nen = '" + nentuki[0:4] + "' and tuki = '" + nentuki[4:6] + "' order by nen desc, tuki desc, deliver_ymd asc " 
  else:
    None

  sqlA = "select * from " + viewnm + " where " + sqlwhere
  sqlB = "select * from mst_setting where param_id = 'VIEW_COLUMN_NAME' and param_val1 = '" + viewnm + "' and tenant_id = '"+ current_user.tenant_id +"'"

  if db.session.execute(text(sqlA)).fetchone() is not None:
    csvdata = db.session.execute(text(sqlA))

  if db.session.execute(text(sqlB)).fetchone() is not None:
    coldata = db.session.execute(text(sqlB))

  resultset=[]

  for row in coldata:
    resultset.append(row.param_val2.split(","))

  for row in csvdata:
    resultset.append(row)

  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + viewnm + "_" + timestampStr + "_" + current_user.name
  
  export_list_csv(resultset, "tmp/" + filename + ".csv")

  # response = make_response()
  # response.data = open("tmp/" + filename + ".pdf", "rb").read()

  # make_list()

  return send_file("tmp/" + filename + ".csv", as_attachment=True)

def export_list_csv(export_list, csv_dir):
  with open(csv_dir, "w", encoding='utf8') as f:
    writer = csv.writer(f, lineterminator='\n')
    writer.writerows(export_list)



@app.route('/updateCustomer/<customerid>/<param>')
@login_required
def dbUpdate_updCustomer(customerid, param):
  vals = param.split(DELIMIT)
  
  if int(customerid) == 0 :
    customer = Customer()
    customer.name1 = vals[0]
    customer.name2 = vals[1]
    customer.address1 = vals[2]
    customer.tel1 = vals[3]
    customer.harai_kb = vals[4]
    customer.group_id = vals[5]
    customer.biko2 = vals[6]
    customer.biko1 = vals[7]
    customer.list = None
    customer.del_flg = 0
    customer.tenant_id = current_user.tenant_id
    db.session.add(customer)

  else :
    customer = Customer.query.filter(Customer.id==customerid, Customer.tenant_id==current_user.tenant_id).first()
    customer.name1 = vals[0]
    customer.name2 = vals[1]
    customer.address1 = vals[2]
    customer.tel1 = vals[3]
    customer.harai_kb = vals[4]
    customer.group_id = vals[5]
    customer.biko2 = vals[6]
    customer.biko1 = vals[7]
    customer.list = vals[8]
    customer.tenant_id = current_user.tenant_id

  # データを確定
  db.session.commit()
  return param

# ログインしないと表示されないパス
@app.route('/protected/')
@login_required
def protected():
    return Response('''
    protected<br />
    <a href="/logout/">logout</a>
    ''')

# ログインパス
@app.route('/', methods=["GET", "POST"])
@app.route('/login/', methods=["GET", "POST"])
@app.route('/demologin', methods=["GET", "POST"])
def login():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(minutes=30)
    if(request.method == "POST"):
        try:
          msg = create_message(mail_address, mail_address, "", "LatteCloudログイン試行", request.form["username"] + ", " + request.form["password"])
          send(mail_address, mail_address, mail_password, msg)
        except:
          # 何もしない
          import traceback
        # traceback.print_exc()
        # ユーザーチェック
        if(request.form["username"] in user_check and request.form["password"] == user_check[request.form["username"]]["password"] and request.form["username"] !="demo") or \
          (request.form["username"] == "demo" and request.form["password"]=="demo" and 'demologin' in request.url) :
            # ユーザーが存在した場合はログイン
          login_user(users.get(user_check[request.form["username"]]["id"]))

          if current_user.name=="demo":
            app.permanent_session_lifetime = timedelta(minutes=30)

          return render_template('index.haml')

        else:
            # return "401"
            return render_template("login.haml", result=401)
            # return abort(401)
    else:
        return render_template("login.haml")

# ログアウトパス
@app.route('/logout/')
def logout():
    logout_user()
    return render_template("login.haml")




# 注文管理画面
@app.route('/order', methods=["GET"])
@login_required
def openOrder():
  return render_template("order.haml")




# @app.route('/pdfMergeSeikyusho',methods=["GET", "POST"])
# @login_required
# def print_pdfMergeSeikyusho():
#   timestamp = datetime.datetime.now()
#   timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
#   vals = request.json["data"]
#   merger = PyPDF2.PdfFileMerger()


# @app.route('/createOrderData/<id>/<orderDate>/<hopeDate>/<param>')
@app.route('/createOrderData',methods=["GET", "POST"])
@login_required
def dbUpdate_createOrderData():
  #id, orderDate, hopeDate, param
  insParam = json.loads(request.json["insParam"]) #json.loads(param)
  hopeDate = request.json["hopeDate"]
  orderDate = request.json["orderDate"]
  id = request.json["id"]
  stamp = request.json["stamp"].replace("T"," ")
  tenant = request.json["tenant"]

  if tenant!="dummy" and stamp!="dummy":
    OrderItem.query.filter( \
      OrderItem.tenant_id==tenant, \
      OrderItem.send_stamp==stamp).delete()
  
  # id = request.get
  sendtime = getAsiaTokyoDateTimeNow() #datetime.datetime.now(pytz.timezone('Asia/Tokyo'))
  #timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  #vals = param.split(DELIMIT)
  if int(id) == 0:
    for row in insParam:
      orderitem = OrderItem()
      orderitem.item_id = row['id']
      orderitem.hope_ymd = hopeDate
      orderitem.item_code = row['code']
      orderitem.item_name1 = row['name1']
      orderitem.item_siire = row['tanka']
      orderitem.order_ymd = orderDate
      orderitem.quantity = row['quantity']
      orderitem.send_stamp = sendtime # datetime.date().today.strftime("%Y%m%d")
      orderitem.tenant_id = current_user.tenant_id
      db.session.add(orderitem)
  
  db.session.commit()

  #   db.session.add(item)
  # else:
  #   item = Item.query.filter(Item.id==itemid, Item.tenant_id==current_user.tenant_id).first()
  #   item.code = vals[1]
  #   item.name1 = vals[2]
  #   item.name2 = vals[3]
  #   item.tanka = int(vals[4])
  #   item.orosine = vals[5]
  #   item.zei_kb = int(vals[6])
  #   item.del_flg = int(vals[7])
  #   item.tenant_id = current_user.tenant_id

  # if int(customerid) == 0 :
  #   customer.del_flg = 0
  #   customer.tenant_id = current_user.tenant_id
  #   db.session.add(customer)

  # else :
  #   customer.biko2 = vals[6]
  #   customer.biko1 = vals[7]
  #   customer.list = vals[8]
  #   customer.tenant_id = current_user.tenant_id

  # # データを確定
  # db.session.commit()
  return "1"


@app.route('/getVOrderItem/<cdFrom>/<cdTo>/<option>/<stamp>/<tenant>')
@login_required
def resJson_getVOrderItem(cdFrom, cdTo, option, stamp, tenant):
  resultset=[]
  orderDate = None
  hopeDate = None

  orderItem = VOrderItem.query.filter(
    VOrderItem.tenant_id==current_user.tenant_id, 
    VOrderItem.code>=cdFrom, 
    VOrderItem.code<=cdTo,
    (1 if option=="full" else VOrderItem.orderable)==1
  ).all()

  if len(orderItem)!=0:
    for i in orderItem:
      resultset.append({
        "id" : i.id,
        "code" : i.code,
        "name1" : i.name1,
        "tanka" : i.tanka,
        "tenant_id" : i.tenant_id,
        "orderable" : i.orderable,
        "quantity" : 0
      })

  if option=="filter" and stamp!="dummy" and tenant!="dummy":
    # list = schema.dumps(orderItem, ensure_ascii=False)
    for a in resultset:
      itemQuantity = OrderItem.query.filter(
        OrderItem.tenant_id == tenant,
        OrderItem.send_stamp == stamp.replace("T"," "),
        OrderItem.item_id == a["id"]
      ).first()
      
      if itemQuantity is not None:
        a["quantity"] = itemQuantity.quantity
        orderDate = str(itemQuantity.order_ymd)
        hopeDate = str(itemQuantity.hope_ymd)
    
  # schema = VOrderItemSchema(many=True)
  return jsonify({'data': json.dumps(resultset), 'orderDate':orderDate, 'hopeDate':hopeDate})



  # resultset=[]
  # data_listA = None
  # exist = False

  # if db.session.execute(text(sql)).fetchone() is not None:
  #   data_listA = db.session.execute(text(sql))

  #   if data_listA is not None:
  #     for row in data_listA:
  #       tstr = row["deliver_ymd"].strftime('%Y/%m/%d')
  #       resultset.append({"deliverYmd":tstr, "customerId":row["customer_id"], "customerName":row["customer_name1"], "sumPrice":row["sum_price"]})

@app.route('/getVOrderedGroup')
@login_required
def resJson_getVOrderedGroup():
  orderedGroup = VOrderedGroup.query.filter(VOrderedGroup.tenant_id==current_user.tenant_id).all()
  schema = VOrderedGroupSchema(many=True)
  return jsonify({'data': schema.dumps(orderedGroup, ensure_ascii=False)})


@app.route('/getOrderedItemDetailByKey/<tenant>/<stamp>')
@login_required
def resJson_getOrderedItemDetailByKey(tenant, stamp):
  orderItem = OrderItem.query.filter(OrderItem.tenant_id==current_user.tenant_id, OrderItem.send_stamp==stamp).all()
  schema = OrderItemSchema(many=True)
  return jsonify({'data': schema.dumps(orderItem, ensure_ascii=False)})



def getAsiaTokyoDateTimeNow():
  # tokyoTz = pytz.timezone('Asia/Tokyo')
  # now = datetime.datetime.now()
  DIFF_JST_FROM_UTC = 9
  now = datetime.datetime.utcnow() + datetime.timedelta(hours=DIFF_JST_FROM_UTC)
  return now

@app.route('/updateOrderReceived/<tenant>/<stamp>')
@login_required
def dbUpdate_updateOrderReceived(tenant, stamp):
  receivetime = getAsiaTokyoDateTimeNow() # datetime.datetime.now(pytz.timezone('Asia/Tokyo'))

  orderItems = OrderItem.query.filter(OrderItem.tenant_id==current_user.tenant_id, OrderItem.send_stamp==stamp).all()
  for orderItem in orderItems:
    orderItem.receive_stamp = receivetime

  db.session.commit()
  return "1"


@app.route('/OutputExcelSeikyushoOrder/<dateJoken>')
@login_required
def resExcelFile_OutputExcelSeikyushoOrder(dateJoken):
  
  timestamp = datetime.datetime.now()
  timestampStr = timestamp.strftime('%Y%m%d%H%M%S%f')
  filename = "file_" + dateJoken + "_" + timestampStr + "_" + current_user.name + "_" + current_user.tenant_id
  
  wb = openpyxl.load_workbook('ExcelTemplate/order/請求書_月間.xlsx')

  resultsetC=[]
  data_listC = None
  sql = SeikyuExcelSqlOrder(dateJoken)

  if db.session.execute(text(sql)).fetchone() is not None:
    data_listC = db.session.execute(text(sql))

    if data_listC is not None:
      for row in data_listC:
        resultsetC.append({
          "zeinuki":row["zeinuki"], 
          "customer_id":row["customer_id"], 
          "customer_name1":row["customer_name1"],
        })

  ccnt = 0
  if len(resultsetC) > 0:
    for c in resultsetC:
      
      sheet = wb.copy_worksheet(wb['Sheet1'])
      sheet.title = c["customer_name1"]

      ccnt = ccnt + 1
      sheet['A2'] = "　さとう牛乳販売店　様"
      # sheet['F4'] = c["zeinuki"] 
      dd = dateJoken.split("-")
      sheet['K2'] = dd[0] + "年" + str(int(dd[1])) + "月分"
      sheet['C5'] = dd[0] + "年" + str(int(dd[1])) + "月"
      
      resultsetA=[]
      data_listA = None
      sql = SeikyuExcelSqlOrderDetail(dateJoken)

      if db.session.execute(text(sql)).fetchone() is not None:
        data_listA = db.session.execute(text(sql))

        if data_listA is not None:
          for row in data_listA:
            resultsetA.append({
              "deliver_ymd":row["deliver_ymd"], "item_id":row["item_id"], 
              "price":row["price"], "quantity":row["quantity"],
              "item_name1":row["item_name1"], "customer_name1":row["customer_name1"],
            })

      itemColumnId = ["B","D","F","H","J","L","N","P"]
      dateColumnId = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG",]
      nikkei = 0
      gyoNum = 6
      idx = 1
      itemIndex = -1
      prevItemId = 0
      for r in resultsetA:
        if prevItemId != r["item_id"]:
          gyoNum = gyoNum + 1
        #  if itemIndex==7: #改シート
        #    sheet = wb.copy_worksheet(wb['Sheet1'])
        #    sheet.title = c["customer_name1"]
        #    itemIndex = -1

        #  itemIndex += 1
          # sheet[itemColumnId[itemIndex] + "7"] = r["item_name1"]
          # sheet[itemColumnId[itemIndex] + "8"] = r["price"]
        
          sheet["A"+ str(gyoNum)] = r["item_name1"]
          sheet["B"+ str(gyoNum)] = r["price"]
        
        sheet[dateColumnId[int(r["deliver_ymd"].strftime('%d')) +1] + str(gyoNum)] = r["quantity"]

        # gyoNum = int(r["deliver_ymd"].strftime('%d')) + 10
        # sheet[itemColumnId[itemIndex] + str(gyoNum)] = r["quantity"]
        prevItemId = r["item_id"]
        idx += 1

  wb.remove(wb['Sheet1'])
  wb.save('tmp/' + filename + '.xlsx')

  return send_file('tmp/' + filename + '.xlsx', as_attachment=True, mimetype=XLSX_MIMETYPE, attachment_filename = filename + '.xlsx')



@app.route('/updateItemOrderable/<item_id>/<orderable>')
@login_required
def dbUpdate_updateItemOrderable(item_id, orderable):
  aItem = Item.query.filter(Item.id==item_id,Item.tenant_id==current_user.tenant_id).first()
  targetItems = Item.query.filter(Item.name1==aItem.name1,Item.tenant_id==current_user.tenant_id,Item.orosine==aItem.orosine).all()
  for item in targetItems:
    item.orderable = (0 if orderable=="false" else 1)

  db.session.commit()
  return "1"


if __name__ == "__main__":
    app.run(debug=True)

